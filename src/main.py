"""
Главный модуль приложения.
Содержит функционал для работы с банковскими транзакциями и другие утилиты.
"""

import json
from typing import Any, Dict, List

from src.decorators.log_decorator import log
from src.processing import (
    filter_by_state,
    search_transactions_by_description,
    sort_by_date,
    read_csv_transactions,      # Добавляем импорт
    read_json_transactions,     # Добавляем импорт
)
from src.text_utils import reverse_text

# ============ СУЩЕСТВУЮЩИЙ ФУНКЦИОНАЛ ============

@log()
def test_function():
    """Тестовая функция для проверки декоратора."""
    print("Тестовая функция")

def run_text_reverser():
    """Запускает функционал реверса текста."""
    text = input("Введите любой текст для реверса: ")
    print(reverse_text(text))

# ============ НОВЫЙ ФУНКЦИОНАЛ ДЛЯ ДЗ 13.2 ============

def load_transactions_from_json(filepath: str) -> List[Dict[str, Any]]:
    """
    Загружает транзакции из JSON-файла.
    Использует функцию из processing.py
    """
    transactions = read_json_transactions(filepath)
    if not transactions:
        print(f"Файл {filepath} не найден или пуст.")
    return transactions

def load_transactions_from_csv(filepath: str) -> List[Dict[str, Any]]:
    """
    Загружает транзакции из CSV-файла.
    Использует функцию из processing.py
    """
    transactions = read_csv_transactions(filepath)
    if not transactions:
        print(f"Файл {filepath} не найден или пуст.")
    return transactions

def load_transactions_from_xlsx(filepath: str) -> List[Dict[str, Any]]:
    """
    Загружает транзакции из XLSX-файла.
    """
    try:
        from openpyxl import load_workbook
        transactions = []
        wb = load_workbook(filename=filepath, data_only=True)
        ws = wb.active

        # Получаем заголовки из первой строки
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # Читаем данные
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and any(cell is not None for cell in row):
                transaction = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        transaction[header] = row[i]
                transactions.append(transaction)

        wb.close()
        return transactions
    except ImportError:
        print("Библиотека openpyxl не установлена. Установите: pip install openpyxl")
        return []
    except Exception as e:
        print(f"Ошибка загрузки XLSX: {e}")
        return []

def generate_test_transactions() -> List[Dict[str, Any]]:
    """
    Генерирует тестовый набор транзакций для демонстрации.
    """
    return [
        {
            "id": 1,
            "date": "2019-12-08",
            "description": "Открытие вклада",
            "amount": 40542,
            "currency": "руб.",
            "state": "EXECUTED",
        },
        {
            "id": 2,
            "date": "2019-11-12",
            "description": "Перевод с карты на карту",
            "amount": 130,
            "currency": "USD",
            "state": "EXECUTED",
        },
        {
            "id": 3,
            "date": "2018-07-18",
            "description": "Перевод организации",
            "amount": 8390,
            "currency": "руб.",
            "state": "CANCELED",
        },
        {
            "id": 4,
            "date": "2018-06-03",
            "description": "Перевод со счета на счет",
            "amount": 8200,
            "currency": "EUR",
            "state": "PENDING",
        },
        {
            "id": 5,
            "date": "2020-01-01",
            "description": "Оплата интернета",
            "amount": 500,
            "currency": "руб.",
            "state": "EXECUTED",
        },
    ]

def filter_ruble_transactions(
    transactions: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """
    Оставляет только рублевые транзакции.
    """
    return [tx for tx in transactions if tx.get("currency", "").lower() == "руб."]

def print_transactions(transactions: List[Dict[str, Any]]) -> None:
    """
    Красиво выводит список транзакций в консоль.
    """
    if not transactions:
        print("\nНе найдено ни одной транзакции, подходящей под ваши условия фильтрации")
        return

    print(f"\nВсего банковских операций в выборке: {len(transactions)}")
    print("-" * 60)

    for tx in transactions:
        date = tx.get("date", "Дата не указана")
        desc = tx.get("description", "Без описания")
        amount = tx.get("amount", 0)
        currency = tx.get("currency", "")
        print(f"{date} {desc}")
        print(f"Сумма: {amount} {currency}")
        print("-" * 60)

def run_bank_processor():
    """
    Запускает основной функционал программы для работы с банковскими транзакциями.
    """
    print("Привет! Добро пожаловать в программу работы с банковскими транзакциями.")
    print("Выберите необходимый пункт меню:")
    print("1. Получить информацию о транзакциях из JSON-файла")
    print("2. Получить информацию о транзакциях из CSV-файла")
    print("3. Получить информацию о транзакциях из XLSX-файла")

    choice = input("\nВаш выбор: ").strip()
    transactions: List[Dict[str, Any]] = []

    if choice == "1":
        print("Для обработки выбран JSON-файл.")
        filepath = input("Введите путь к JSON-файлу (или нажмите Enter для тестовых данных): ").strip()
        if filepath:
            transactions = load_transactions_from_json(filepath)
        if not transactions:
            print("Файл не найден или пуст. Загружаю тестовые данные.")
            transactions = generate_test_transactions()
    elif choice == "2":
        print("Для обработки выбран CSV-файл.")
        filepath = input("Введите путь к CSV-файлу (или нажмите Enter для тестовых данных): ").strip()
        if filepath:
            transactions = load_transactions_from_csv(filepath)
        if not transactions:
            print("Файл не найден или пуст. Загружаю тестовые данные.")
            transactions = generate_test_transactions()
    elif choice == "3":
        print("Для обработки выбран XLSX-файл.")
        filepath = input("Введите путь к XLSX-файлу (или нажмите Enter для тестовых данных): ").strip()
        if filepath:
            transactions = load_transactions_from_xlsx(filepath)
        if not transactions:
            print("Файл не найден или пуст. Загружаю тестовые данные.")
            transactions = generate_test_transactions()
    else:
        print("Неверный выбор. Загружаю тестовые данные.")
        transactions = generate_test_transactions()

    if not transactions:
        print("Не удалось загрузить транзакции.")
        return

    # Фильтрация по статусу
    valid_statuses = ["EXECUTED", "CANCELED", "PENDING"]
    while True:
        status_input = input(
            "\nВведите статус, по которому необходимо выполнить фильтрацию.\n"
            "Доступные для фильтровки статусы: EXECUTED, CANCELED, PENDING\n"
            "Ваш статус: "
        ).strip().upper()

        if status_input in valid_statuses:
            transactions = filter_by_state(transactions, status_input)
            print(f'Операции отфильтрованы по статусу "{status_input}"')
            break
        else:
            print(f'Статус операции "{status_input}" недоступен.')

    if not transactions:
        print("Не найдено ни одной транзакции с таким статусом.")
        return

    # Сортировка по дате
    sort_choice = input("\nОтсортировать операции по дате? Да/Нет: ").strip().lower()
    if sort_choice in ["да", "yes", "y", "д"]:
        order = input("Отсортировать по возрастанию или по убыванию? ").strip().lower()
        ascending = order in ["по возрастанию", "возрастанию", "asc", "возраст"]
        transactions = sort_by_date(transactions, ascending)

    # Фильтр по рублям
    ruble_choice = input("\nВыводить только рублевые транзакции? Да/Нет: ").strip().lower()
    if ruble_choice in ["да", "yes", "y", "д"]:
        transactions = filter_ruble_transactions(transactions)

    if not transactions:
        print("Не найдено ни одной транзакции, подходящей под ваши условия фильтрации")
        return

    # Поиск по описанию (НОВАЯ ФУНКЦИЯ!)
    search_choice = input(
        "\nОтфильтровать список транзакций по определенному слову в описании? Да/Нет: "
    ).strip().lower()
    if search_choice in ["да", "yes", "y", "д"]:
        search_word = input("Введите слово для поиска: ").strip()
        if search_word:
            transactions = search_transactions_by_description(transactions, search_word)

    # Вывод результата
    print("\nРаспечатываю итоговый список транзакций...")
    print_transactions(transactions)

def main():
    """
    Основная функция программы.
    Предлагает пользователю выбрать режим работы.
    """
    print("Добро пожаловать в программу!")
    print("Выберите режим работы:")
    print("1. Работа с банковскими транзакциями")
    print("2. Реверс текста (тестовый режим)")

    mode = input("\nВаш выбор: ").strip()

    if mode == "1":
        run_bank_processor()
    elif mode == "2":
        test_function()
        run_text_reverser()
    else:
        print("Неверный выбор. Запускаю режим работы с банковскими транзакциями.")
        run_bank_processor()

if __name__ == "__main__":
    main()
